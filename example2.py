import openpyxl
import os
import logging

print(os.name)

logger = logging.getLogger("MAIN")
logger.error("My error")

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# calculate products per suplier
products_per_supplier = {}
print(product_list.max_row)  # max_row returns number of rows in a spreadsheet
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        print(f"adding a new supplier, {supplier_name}")
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)

# calculate total value of inventory per suplier
total_value_per_supplier = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

print(total_value_per_supplier)

# calculate products with inventory < 10
products_with_inventory_under_10 = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    products_num = int(product_list.cell(product_row, 1).value)

    if inventory < 10:
        products_with_inventory_under_10[products_num] = int(inventory)

print(products_with_inventory_under_10)  # product number: inventory items

# add a column
products_with_inventory_under_10 = {}
for product_row in range(2, product_list.max_row + 1):
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    inventory_price = product_list.cell(product_row, 5)
    inventory_price.value = inventory * price


inv_file.save("inv_with_total_value.xlsx")  # new file is created
